import openpyxl
import pandas as pd

wb = openpyxl.load_workbook('datos_proy2.xlsx')
ws = wb['solicitudes de marcas']
# This will fail if there is no hyperlink to target
toFormat = f'{ws.cell(row=2, column=7).value}'
hyperlink = toFormat.split('"')[1]
print(hyperlink)

# ELIMINAR INSTANCIA 119 PORQUE ESTÁ REPETIDA


def xlsx_to_sql(sql_file, table_name):
    
    columnas_deseadas = ["tipo de signo", "Signo", "Descripción etiqueta"]

    df = pd.read_excel("datos_proy2.xlsx", "solicitudes de marcas", usecols=columnas_deseadas)

    cols_list = df.columns.tolist()
    id_sd = 1 #clave postiza a cada signo

    for i in range(len(cols_list)):
        cols_list[i] = cols_list[i].replace(" ","_")
        cols_list[i] = cols_list[i].lower()
        if (cols_list[i] == "tipo_de_signo"):
            cols_list[i] == "tipo"
        elif (cols_list[i] == "signo"):
            cols_list[i] = "nombre"
        elif (cols_list[i] == "descripción_etiqueta"):
            cols_list[i] = "descripción"
    cols_list.append("imagen_correspondiente")
    cols_list.append("id_sd")
    
    with open(sql_file, 'w') as f:
        f.write(f"INSERT INTO {table_name} (")
        f.write(", ".join(cols_list))
        f.write(") VALUES \n")

        for i, row in enumerate(df.iterrows()):
            f.write("(")
            #if (i == 4): print(row[1].iloc[0])
            toFormat = f'{ws.cell(row=i+2, column=7).value}'

            toJoin = ", ".join([f"'{str(val)}'" for val in row[1].values])

            if (row[1].iloc[0] == "denominativa"):
                toJoin = toJoin +  ', NULL' + f', {id_sd}'
            else:
                hyperlink = toFormat.split('"')[1]
                toJoin = toJoin + ", '" + hyperlink + "'"
                toJoin = toJoin + f", {id_sd}"
            
            f.write(toJoin)

            id_sd = id_sd + 1

            if i == len(df) - 1:
                f.write(");\n")
            else:
                f.write("),\n")


if __name__ == "__main__":
    sql_file = "proy2_insert_signo_w_link.sql"
    table_name = "signo_distintivo"
    xlsx_to_sql(sql_file, table_name)